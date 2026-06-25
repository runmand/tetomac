import io, re, time, warnings, os, threading
from flask import Flask, request, send_file, jsonify
import psycopg2, psycopg2.extras
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from playwright.sync_api import sync_playwright

warnings.filterwarnings("ignore")

app = Flask(__name__)
URL_SISMAC   = "https://sismac.saude.gov.br/teto_financeiro_anual"
URL_LISTA    = "https://sismac.saude.gov.br/teto_financeiro_brasil_por_estado_municipio"
TABELA_ID    = "tetoFinanceiroBrasil"
FORM_ID      = "formTemplate"
DATABASE_URL = os.environ.get("DATABASE_URL", "")

def get_conn():
    return psycopg2.connect(DATABASE_URL)

# ── INIT BANCO ──

def init_db():
    if not DATABASE_URL:
        print("⚠️  DATABASE_URL não configurada"); return
    try:
        conn = get_conn()
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS localidades (
                    id SERIAL PRIMARY KEY, regiao TEXT, uf TEXT,
                    cod_ibge TEXT, nome TEXT, cod_gestao TEXT,
                    desc_gestao TEXT, teto_atual NUMERIC,
                    atualizado TIMESTAMP DEFAULT NOW(),
                    UNIQUE(cod_ibge, cod_gestao)
                );
                CREATE INDEX IF NOT EXISTS idx_uf   ON localidades(uf);
                CREATE INDEX IF NOT EXISTS idx_nome ON localidades(nome);
                CREATE TABLE IF NOT EXISTS historico (
                    id SERIAL PRIMARY KEY, cod_ibge TEXT, cod_gestao TEXT,
                    ano INTEGER, valor_total NUMERIC, var_valor NUMERIC, var_pct NUMERIC,
                    coletado_em TIMESTAMP DEFAULT NOW(),
                    UNIQUE(cod_ibge, cod_gestao, ano)
                );
                CREATE INDEX IF NOT EXISTS idx_hist ON historico(cod_ibge, cod_gestao);
            """)
            conn.commit()
            cur.execute("SELECT COUNT(*) FROM localidades")
            total_loc = cur.fetchone()[0]
            cur.execute("SELECT COUNT(DISTINCT cod_ibge||cod_gestao) FROM historico WHERE ano > 0")
            total_hist = cur.fetchone()[0]
        conn.close()
        print(f"✅ Banco OK — {total_loc} localidades | {total_hist} com histórico")
        if total_loc == 0:
            print("📥 Populando localidades em background...")
            threading.Thread(target=popular_localidades, daemon=True).start()
    except Exception as e:
        print(f"❌ Erro banco: {e}")

def popular_localidades():
    import requests
    from bs4 import BeautifulSoup
    print("🌐 Coletando municípios do SISMAC...")
    headers = {"User-Agent": "Mozilla/5.0 Chrome/124.0", "Accept-Language": "pt-BR"}
    session = requests.Session(); session.headers.update(headers)
    try: session.get("https://sismac.saude.gov.br/inicio", timeout=15)
    except: pass

    muns = {}
    def _cdata(xml):
        m = re.search(r"<!\[CDATA\[(.*?)\]\]>", xml, re.DOTALL)
        return m.group(1) if m else xml
    def _brl(txt):
        t = str(txt).strip().replace(".", "").replace(",", ".")
        try: return float(t)
        except: return 0.0
    def _add(html):
        from bs4 import BeautifulSoup
        for tr in BeautifulSoup(html, "lxml").find_all("tr"):
            spans = tr.find_all("span", title=True)
            vals = [s["title"] for s in spans[:7]] if len(spans)>=7 else [td.get_text(strip=True) for td in tr.find_all("td")[:7]]
            if len(vals)<7: continue
            r,uf,cod,nome,cg,dg,teto = vals
            if not re.match(r"^\d{6}$",cod): continue
            muns[cod+"_"+cg] = {"regiao":r,"uf":uf,"cod_ibge":cod,"nome":nome,"cod_gestao":cg,"desc_gestao":dg,"teto_atual":_brl(teto)}

    try:
        r = session.get(URL_LISTA, timeout=30)
        from bs4 import BeautifulSoup
        vs_inp = BeautifulSoup(r.text,"lxml").find("input",{"name":"javax.faces.ViewState"})
        vs = vs_inp["value"] if vs_inp else ""
        _add(r.text); print(f"   P1: {len(muns)}")
        r2 = session.post(URL_LISTA, timeout=60, data={
            "javax.faces.partial.ajax":"true","javax.faces.source":TABELA_ID,
            "javax.faces.partial.execute":TABELA_ID,"javax.faces.partial.render":TABELA_ID,
            f"{TABELA_ID}_encodeFeature":"true",f"{TABELA_ID}_rppDD":"1000",
            FORM_ID:FORM_ID,"javax.faces.ViewState":vs})
        _add(_cdata(r2.text))
        sv = BeautifulSoup(r2.text,"lxml").find("input",{"name":"javax.faces.ViewState"})
        if sv: vs = sv["value"]
        print(f"   1000/pág: {len(muns)}")
        pagina=2; sem_nov=0
        while pagina<=100:
            resp = session.post(URL_LISTA, timeout=30, data={
                "javax.faces.partial.ajax":"true","javax.faces.source":TABELA_ID,
                "javax.faces.partial.execute":TABELA_ID,"javax.faces.partial.render":TABELA_ID,
                f"{TABELA_ID}_pagination":"true",f"{TABELA_ID}_first":str((pagina-1)*1000),
                f"{TABELA_ID}_rows":"1000",f"{TABELA_ID}_skipChildren":"true",
                f"{TABELA_ID}_encodeFeature":"true",FORM_ID:FORM_ID,"javax.faces.ViewState":vs})
            antes=len(muns); _add(_cdata(resp.text)); novos=len(muns)-antes
            sv=BeautifulSoup(resp.text,"lxml").find("input",{"name":"javax.faces.ViewState"})
            if sv: vs=sv["value"]
            sem_nov=0 if novos else sem_nov+1
            if sem_nov>=2: break
            pagina+=1; time.sleep(0.3)
    except Exception as e:
        print(f"⚠️ Erro coleta: {e}")

    if not muns: return
    conn = get_conn()
    with conn.cursor() as cur:
        for m in muns.values():
            cur.execute("""INSERT INTO localidades (regiao,uf,cod_ibge,nome,cod_gestao,desc_gestao,teto_atual)
                VALUES (%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (cod_ibge,cod_gestao) DO UPDATE SET teto_atual=EXCLUDED.teto_atual,atualizado=NOW()""",
                (m["regiao"],m["uf"],m["cod_ibge"],m["nome"],m["cod_gestao"],m["desc_gestao"],m["teto_atual"]))
    conn.commit(); conn.close()
    print(f"✅ {len(muns)} localidades salvas!")

# ── ESTILOS EXCEL ──

COR_H="D5DDE5"; COR_P="EEF2F5"; COR_I="FFFFFF"; COR_T="1A2A3A"
def _borda():
    s=Side(style="thin",color="B5C3CC"); return Border(left=s,right=s,top=s,bottom=s)
def _header(ws,row,cols):
    fill=PatternFill("solid",fgColor=COR_H); font=Font(name="Arial",bold=True,size=9,color=COR_T)
    for col in range(1,cols+1):
        c=ws.cell(row=row,column=col); c.fill=fill; c.font=font; c.border=_borda()
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def _pct(a,b): return (b-a)/a*100 if a else None
def _parse_num(txt):
    if not txt: return 0.0
    t=re.sub(r"[^\d,\.\-]","",str(txt).strip())
    if "," in t and "." in t: t=t.replace(".","").replace(",",".")
    elif "," in t: t=t.replace(",",".")
    try: return float(t)
    except: return 0.0

# ── BUSCA NO BANCO ──

def buscar_no_banco(cod_ibge, cod_gestao, anos_alvo):
    conn = get_conn()
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        query = """SELECT ano, valor_total, var_valor as variacao_valor, var_pct as variacao_pct
                   FROM historico WHERE cod_ibge=%s AND cod_gestao=%s AND ano > 0"""
        params = [cod_ibge, cod_gestao]
        if anos_alvo:
            query += " AND ano = ANY(%s)"
            params.append(list(anos_alvo))
        query += " ORDER BY ano"
        cur.execute(query, params)
        rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

def buscar_localidade(nome, aba, uf=""):
    """Busca localidade priorizando match exato de UF, depois quem tem histórico."""
    conn = get_conn()
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        if aba == "Estado":
            # 1. Match exato de UF com histórico
            cur.execute("""
                SELECT l.cod_ibge, l.cod_gestao, l.nome FROM localidades l
                WHERE l.uf = %s
                AND (l.cod_ibge || l.cod_gestao) IN (
                    SELECT cod_ibge || cod_gestao FROM historico WHERE ano > 0
                )
                LIMIT 1
            """, (nome.upper(),))
            row = cur.fetchone()
            if not row:
                # 2. Match por nome com histórico
                cur.execute("""
                    SELECT l.cod_ibge, l.cod_gestao, l.nome FROM localidades l
                    WHERE l.nome ILIKE %s
                    AND (l.cod_ibge || l.cod_gestao) IN (
                        SELECT cod_ibge || cod_gestao FROM historico WHERE ano > 0
                    )
                    LIMIT 1
                """, (f"%{nome}%",))
                row = cur.fetchone()
            if not row:
                # 3. Fallback sem histórico
                cur.execute("""SELECT cod_ibge, cod_gestao, nome FROM localidades
                              WHERE uf = %s AND desc_gestao IN ('Total UF','Gestão Estadual')
                              ORDER BY desc_gestao DESC LIMIT 1""", (nome.upper(),))
                row = cur.fetchone()
        else:
            # Município — busca com prioridade: exato+UF > parcial+UF > exato sem UF > parcial sem UF
            row = None
            nome_upper = nome.upper()
            uf_upper = uf.upper() if uf else ""

            if uf_upper:
                # 1. Match exato por nome + UF
                cur.execute("""
                    SELECT l.cod_ibge, l.cod_gestao, l.nome FROM localidades l
                    WHERE UPPER(l.nome) = %s AND l.uf = %s AND l.desc_gestao='Gestão Municipal'
                    AND (l.cod_ibge || l.cod_gestao) IN (
                        SELECT cod_ibge || cod_gestao FROM historico WHERE ano > 0
                    )
                    LIMIT 1
                """, (nome_upper, uf_upper))
                row = cur.fetchone()

            if not row and uf_upper:
                # 2. Match parcial por nome + UF
                cur.execute("""
                    SELECT l.cod_ibge, l.cod_gestao, l.nome FROM localidades l
                    WHERE l.nome ILIKE %s AND l.uf = %s AND l.desc_gestao='Gestão Municipal'
                    AND (l.cod_ibge || l.cod_gestao) IN (
                        SELECT cod_ibge || cod_gestao FROM historico WHERE ano > 0
                    )
                    ORDER BY LENGTH(l.nome) ASC
                    LIMIT 1
                """, (f"%{nome}%", uf_upper))
                row = cur.fetchone()

            if not row:
                # 3. Match exato sem UF
                cur.execute("""
                    SELECT l.cod_ibge, l.cod_gestao, l.nome FROM localidades l
                    WHERE UPPER(l.nome) = %s AND l.desc_gestao='Gestão Municipal'
                    AND (l.cod_ibge || l.cod_gestao) IN (
                        SELECT cod_ibge || cod_gestao FROM historico WHERE ano > 0
                    )
                    LIMIT 1
                """, (nome_upper,))
                row = cur.fetchone()
            if not row:
                cur.execute("""SELECT cod_ibge, cod_gestao, nome FROM localidades
                              WHERE nome ILIKE %s AND desc_gestao='Gestão Municipal'
                              LIMIT 1""", (f"%{nome}%",))
                row = cur.fetchone()
    conn.close()
    return dict(row) if row else None
def _parse_excel_bytes(conteudo, anos_alvo):
    wb=openpyxl.load_workbook(io.BytesIO(conteudo)); ws=wb.active
    rows=list(ws.iter_rows(values_only=True))
    hi=next((i for i,r in enumerate(rows) if any(r)),0)
    dados=[]
    for row in rows[hi+1:]:
        if not row or not row[1]: continue
        try: ano=int(float(str(row[1])))
        except: continue
        if ano<2000: continue
        if anos_alvo and ano not in anos_alvo: continue
        dados.append({"ano":ano,"valor_total":float(row[8] or 0),
                      "variacao_valor":float(row[9] or 0),"variacao_pct":float(row[10] or 0)})
    dados.sort(key=lambda x:x["ano"]); return dados

def _ler_html(page, anos_alvo):
    dados=[]
    for tr in page.query_selector_all("table tbody tr"):
        tds=tr.query_selector_all("td")
        if len(tds)<3: continue
        textos=[td.inner_text().strip() for td in tds]
        try: ano=int(re.sub(r"\D","",textos[0]))
        except: continue
        if ano<2000 or ano>2100: continue
        if anos_alvo and ano not in anos_alvo: continue
        if len(textos)>=10: vt,vv,vp=_parse_num(textos[7]),_parse_num(textos[8]),_parse_num(textos[9])
        elif len(textos)>=4: vt,vv,vp=_parse_num(textos[1]),_parse_num(textos[2]),_parse_num(textos[3])
        else: continue
        dados.append({"ano":ano,"valor_total":vt,"variacao_valor":vv,"variacao_pct":vp})
    dados.sort(key=lambda x:x["ano"]); return dados

def coletar_sismac(busca, aba, anos_alvo):
    with sync_playwright() as pw:
        browser=pw.chromium.launch(headless=True,args=[
            "--no-sandbox","--disable-setuid-sandbox","--disable-dev-shm-usage",
            "--disable-gpu","--disable-software-rasterizer","--single-process"])
        context=browser.new_context(accept_downloads=True)
        page=context.new_page(); page.on("dialog",lambda d:d.accept())
        page.goto(URL_SISMAC,wait_until="networkidle",timeout=90000); time.sleep(3)
        try: page.get_by_role("tab",name=aba).click(); time.sleep(1)
        except:
            try: page.click(f"text={aba}"); time.sleep(1)
            except: pass
        campo=page.locator("input.ui-autocomplete-input").first
        campo.wait_for(timeout=15000); campo.click(); campo.fill(""); time.sleep(0.5)
        campo.type(busca[:5],delay=150); time.sleep(3)
        page.wait_for_selector(".ui-autocomplete-item",timeout=10000)
        sugestoes=page.locator(".ui-autocomplete-item").all()
        for s in sugestoes:
            if busca.upper().split()[0] in s.inner_text().strip().upper(): s.click(); break
        else:
            if sugestoes: sugestoes[0].click()
        time.sleep(1)
        for sel in ["button[id*='pesquis']","span[id*='pesquis']","a[id*='pesquis']",".fa-search"]:
            btn=page.query_selector(sel)
            if btn: btn.click(); break
        else: page.keyboard.press("Enter")
        page.wait_for_load_state("networkidle",timeout=40000); time.sleep(5)
        try:
            with page.expect_download(timeout=8000) as dl:
                btn=page.query_selector("a img[src*='excel'],img[src*='excel'],a[href*='excel'],img[src*='xls']")
                if not btn: raise Exception("sem botão")
                btn.click()
            from pathlib import Path
            dados=_parse_excel_bytes(Path(dl.value.path()).read_bytes(),anos_alvo)
        except: dados=_ler_html(page,anos_alvo)
        browser.close()
    return dados

# ── GERA EXCEL ──

def gerar_excel(nome, dados):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Evolução Teto MAC"
    ws.column_dimensions["A"].width=8; ws.column_dimensions["B"].width=30
    ws.column_dimensions["C"].width=30; ws.column_dimensions["D"].width=14
    row=1
    c=ws.cell(row=row,column=1,value=f"EVOLUÇÃO DO TETO MAC – {nome.upper()}")
    c.font=Font(name="Arial",bold=True,size=11,color=COR_T)
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
    ws.row_dimensions[row].height=22; row+=1
    for col,h in enumerate(["ANO","VALOR TOTAL – TETO MAC (R$)","VALOR DA VARIAÇÃO – TETO MAC (R$)","VARIAÇÃO (%)"],1):
        ws.cell(row=row,column=col,value=h)
    _header(ws,row,4); ws.row_dimensions[row].height=28; row+=1
    for i,d in enumerate(dados):
        bg=COR_I if i%2==0 else COR_P; fill=PatternFill("solid",fgColor=bg); fn=Font(name="Arial",size=9)
        c1=ws.cell(row=row+i,column=1,value=str(d["ano"]))
        c1.font=Font(name="Arial",bold=True,size=9); c1.fill=fill; c1.border=_borda()
        c2=ws.cell(row=row+i,column=2,value=float(d.get("valor_total") or 0))
        c2.font=fn; c2.fill=fill; c2.border=_borda(); c2.number_format='#,##0.00'; c2.alignment=Alignment(horizontal="right")
        c3=ws.cell(row=row+i,column=3,value=float(d.get("variacao_valor",d.get("var_valor")) or 0))
        c3.font=fn; c3.fill=fill; c3.border=_borda(); c3.number_format='#,##0.00'; c3.alignment=Alignment(horizontal="right")
        pct=float(d.get("variacao_pct",d.get("var_pct")) or 0)
        c4=ws.cell(row=row+i,column=4,value=pct/100 if pct else None)
        c4.font=fn; c4.fill=fill; c4.border=_borda(); c4.number_format='0.00%'; c4.alignment=Alignment(horizontal="right")
    row+=len(dados)
    if len(dados)>=2:
        vt=_pct(float(dados[0].get("valor_total") or 0),float(dados[-1].get("valor_total") or 0))
        if vt:
            txt=f"VARIAÇÃO {dados[0]['ano']} – {dados[-1]['ano']}: {vt:.2f}%".replace(".",",")
            c=ws.cell(row=row,column=1,value=txt); c.font=Font(name="Arial",bold=True,size=9)
            ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ── ROTAS ──

@app.route("/")
def index():
    return open("index.html",encoding="utf-8").read()

@app.route("/status")
def status():
    try:
        conn=get_conn()
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM localidades"); loc=cur.fetchone()[0]
            cur.execute("SELECT COUNT(DISTINCT cod_ibge||cod_gestao) FROM historico WHERE ano>0"); hist=cur.fetchone()[0]
        conn.close()
        return jsonify({"localidades":loc,"com_historico":hist,"pronto":loc>0})
    except Exception as e:
        return jsonify({"erro":str(e),"pronto":False})

@app.route("/ufs")
def ufs():
    try:
        conn=get_conn()
        with conn.cursor() as cur:
            cur.execute("SELECT DISTINCT uf FROM localidades ORDER BY uf")
            rows=[r[0] for r in cur.fetchall()]
        conn.close(); return jsonify(rows)
    except: return jsonify([])

@app.route("/localidades")
def localidades():
    uf=request.args.get("uf","").strip().upper()
    try:
        conn=get_conn()
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""SELECT uf,cod_ibge,nome,desc_gestao FROM localidades
                          WHERE uf=%s AND desc_gestao='Gestão Municipal' ORDER BY nome""",(uf,))
            rows=cur.fetchall()
        conn.close(); return jsonify([dict(r) for r in rows])
    except: return jsonify([])

@app.route("/buscar", methods=["POST"])
def buscar():
    d=request.json
    busca=d.get("busca","").strip()
    aba=d.get("aba","Estado")
    uf=d.get("uf","").strip()
    anos=d.get("anos",[2022,2023,2024,2025])
    anos_alvo=set(anos) if anos else None
    if not busca: return jsonify({"erro":"Informe o nome"}),400

    # 1. Tenta buscar no banco primeiro
    print(f"🔍 Buscando: '{busca}' aba={aba} uf={uf}")
    loc = buscar_localidade(busca, aba, uf)
    print(f"   Localidade encontrada: {loc}")
    if loc:
        dados = buscar_no_banco(loc["cod_ibge"], loc["cod_gestao"], anos_alvo)
        print(f"   Dados banco: {len(dados)} registros | primeiro: {dados[0] if dados else 'vazio'}")
        if dados:
            nome_real = loc.get("nome", busca)
            print(f"✅ Retornando: {nome_real} — {len(dados)} anos")
            return jsonify({"dados":dados,"nome":nome_real,"fonte":"banco"})

    # 2. Fallback: Playwright
    print(f"🌐 SISMAC: {busca}")
    try:
        dados=coletar_sismac(busca,aba,anos_alvo)
        if not dados: return jsonify({"erro":"Nenhum dado encontrado."}),404
        return jsonify({"dados":dados,"nome":busca,"fonte":"sismac"})
    except Exception as e:
        return jsonify({"erro":str(e)}),500

@app.route("/exportar-estado", methods=["POST"])
def exportar_estado():
    """Gera Excel com todos os municípios do estado + compilado do estado."""
    d = request.json
    uf   = d.get("uf","").strip().upper()
    anos = d.get("anos",[2022,2023,2024,2025])
    anos_alvo = set(anos) if anos else None

    if not uf:
        return jsonify({"erro": "Informe a UF"}), 400

    try:
        conn = get_conn()
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Busca todos os municípios do estado com histórico
            query = """
                SELECT l.nome, l.desc_gestao, l.cod_ibge, l.cod_gestao,
                       h.ano, h.valor_total, h.var_valor, h.var_pct
                FROM localidades l
                JOIN historico h ON h.cod_ibge=l.cod_ibge AND h.cod_gestao=l.cod_gestao
                WHERE l.uf = %s AND h.ano > 0
            """
            params = [uf]
            if anos_alvo:
                query += " AND h.ano = ANY(%s)"
                params.append(list(anos_alvo))
            query += " ORDER BY l.desc_gestao, l.nome, h.ano"
            cur.execute(query, params)
            rows = cur.fetchall()
        conn.close()

        if not rows:
            return jsonify({"erro": "Nenhum dado encontrado"}), 404

        # Gera Excel com uma aba por município + aba compilado
        wb = openpyxl.Workbook()
        ws_comp = wb.active
        ws_comp.title = "Compilado"

        # Cabeçalho compilado
        hdrs = ["Município", "Tipo Gestão", "Ano", "Valor Total (R$)", "Variação (R$)", "Variação (%)"]
        for col, h in enumerate(hdrs, 1):
            c = ws_comp.cell(row=1, column=col, value=h)
            fill = PatternFill("solid", fgColor="D5DDE5")
            font = Font(name="Arial", bold=True, size=9, color="1A2A3A")
            borda = Border(left=Side(style="thin",color="B5C3CC"),right=Side(style="thin",color="B5C3CC"),
                          top=Side(style="thin",color="B5C3CC"),bottom=Side(style="thin",color="B5C3CC"))
            c.fill=fill; c.font=font; c.border=borda
            c.alignment=Alignment(horizontal="center",vertical="center")
        ws_comp.row_dimensions[1].height = 25

        # Agrupa por município
        from collections import defaultdict
        municipios = defaultdict(list)
        for r in rows:
            key = (r["nome"], r["desc_gestao"], r["cod_ibge"], r["cod_gestao"])
            municipios[key].append(r)

        linha = 2
        for (nome, desc, cod_ibge, cod_gestao), dados in municipios.items():
            dados.sort(key=lambda x: x["ano"])
            for i, d in enumerate(dados):
                bg = "FFFFFF" if linha%2==0 else "EEF2F5"
                fill = PatternFill("solid", fgColor=bg)
                fn = Font(name="Arial", size=9)
                borda = Border(left=Side(style="thin",color="B5C3CC"),right=Side(style="thin",color="B5C3CC"),
                              top=Side(style="thin",color="B5C3CC"),bottom=Side(style="thin",color="B5C3CC"))
                vals = [nome, desc, str(d["ano"]), float(d["valor_total"] or 0),
                        float(d["var_valor"] or 0), float(d["var_pct"] or 0)/100 if d["var_pct"] else None]
                for col, v in enumerate(vals, 1):
                    c = ws_comp.cell(row=linha, column=col, value=v)
                    c.fill=fill; c.font=fn; c.border=borda
                    if col == 4: c.number_format='#,##0.00'; c.alignment=Alignment(horizontal="right")
                    if col == 5: c.number_format='#,##0.00'; c.alignment=Alignment(horizontal="right")
                    if col == 6: c.number_format='0.00%'; c.alignment=Alignment(horizontal="right")
                linha += 1

        # Ajusta larguras
        ws_comp.column_dimensions["A"].width = 35
        ws_comp.column_dimensions["B"].width = 18
        ws_comp.column_dimensions["C"].width = 8
        ws_comp.column_dimensions["D"].width = 22
        ws_comp.column_dimensions["E"].width = 22
        ws_comp.column_dimensions["F"].width = 12

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        fname = f"teto_mac_{uf}_completo.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route("/exportar", methods=["POST"])
def exportar():
    d=request.json
    nome=d.get("nome","LOCAL"); dados=d.get("dados",[])
    if not dados: return jsonify({"erro":"Sem dados"}),400
    buf=gerar_excel(nome,dados)
    fname=f"teto_mac_{nome.lower().replace(' ','_')}.xlsx"
    return send_file(buf,as_attachment=True,download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

with app.app_context():
    pass

if __name__=="__main__":
    init_db()
    port=int(os.environ.get("PORT",8000))
    app.run(host="0.0.0.0",port=port)
